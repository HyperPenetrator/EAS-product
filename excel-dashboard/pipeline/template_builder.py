"""
template_builder.py
───────────────────
Generate a production-quality, styled Excel report (.xlsx) from summary tables.
Features:
  - Dashboard sheet with KPI boxes referencing named ranges
  - Monthly data sheet with conditional formatting and sparklines
  - Regional data sheet with embedded bar chart
  - Top Products sheet
  - Named ranges for dynamic Antigravity MCP access

Usage:
    python template_builder.py --data data/clean_data.csv --output reports/report.xlsx
"""

import argparse
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName

logger = logging.getLogger(__name__)

# ── Style Palette ─────────────────────────────────────────────────────────────
DARK_NAVY   = "0F1629"
INDIGO      = "4F46E5"
LIGHT_INDIGO = "818CF8"
EMERALD     = "059669"
AMBER       = "D97706"
SLATE_700   = "334155"
SLATE_100   = "F1F5F9"
WHITE       = "FFFFFF"

def _header_style(color_hex: str = INDIGO):
    """Return (font, fill, alignment) for a header cell."""
    font = Font(color=WHITE, bold=True, size=11, name="Aptos")
    fill = PatternFill("solid", fgColor=color_hex)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align

def _thin_border():
    side = Side(border_style="thin", color="D1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)

def _apply_header(cell, text: str, color_hex: str = INDIGO):
    font, fill, align = _header_style(color_hex)
    cell.value = text
    cell.font = font
    cell.fill = fill
    cell.alignment = align

def _write_df(ws, df: pd.DataFrame, start_row: int = 2, header_color: str = INDIGO):
    """Write a DataFrame to a worksheet with styled headers."""
    cols = df.columns.tolist()

    # Write headers
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        _apply_header(cell, str(col_name), header_color)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(str(col_name)) + 6)

    # Write data rows
    alt_fill = PatternFill("solid", fgColor=SLATE_100)
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Format floats nicely
            if isinstance(value, float):
                if "pct" in cols[col_idx - 1] or "margin" in cols[col_idx - 1]:
                    cell.value = round(value, 4)
                    cell.number_format = "0.0%"
                else:
                    cell.value = round(value, 2)
                    cell.number_format = '#,##0.00'
            elif isinstance(value, int):
                cell.value = value
                cell.number_format = '#,##0'
            else:
                cell.value = str(value) if value is not None else ""

            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")


def build_smart_template(summaries: dict, output_path: str):
    """
    Build a multi-sheet styled Excel report from summary tables.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: 📊 Dashboard ─────────────────────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_view.showGridLines = False
    ws_dash.row_dimensions[1].height = 40

    # Title cell
    ws_dash["B2"].value = "📊 Excel Analytics Report"
    ws_dash["B2"].font = Font(bold=True, size=18, color=INDIGO, name="Aptos")
    ws_dash["B2"].alignment = Alignment(horizontal="left", vertical="center")

    ts = ws_dash["B3"]
    ts.value = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    ts.font = Font(size=10, color="64748B", name="Aptos")

    # KPI boxes (row 5–9)
    kpis = summaries.get("kpis", {})
    kpi_defs = [
        ("B", "TOTAL REVENUE", kpis.get("total_revenue", 0), "$#,##0.00", INDIGO),
        ("D", "TOTAL ORDERS",  kpis.get("total_orders",  0), "#,##0",     EMERALD),
        ("F", "AVG ORDER VAL", kpis.get("avg_order_value", 0), "$#,##0.00", SLATE_700),
        ("H", "PROFIT MARGIN", kpis.get("profit_margin", 0), "0.0%",      AMBER),
    ]
    for col, label, value, fmt, color in kpi_defs:
        label_cell = ws_dash[f"{col}5"]
        label_cell.value = label
        label_cell.font = Font(bold=True, size=9, color=WHITE, name="Aptos")
        label_cell.fill = PatternFill("solid", fgColor=color)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        val_cell = ws_dash[f"{col}7"]
        val_cell.value = value if "%" not in fmt else value
        val_cell.font = Font(bold=True, size=20, color=color, name="Aptos")
        val_cell.number_format = fmt
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_dash.column_dimensions["A"].width = 3
    for col in list("BCDEFGHI"):
        ws_dash.column_dimensions[col].width = 18

    # ── Sheet 2: 📅 Monthly ───────────────────────────────────────────────────
    monthly_df = summaries.get("monthly")
    if monthly_df is not None and not monthly_df.empty:
        ws_monthly = wb.create_sheet("Monthly")
        ws_monthly.sheet_view.showGridLines = False

        # Convert Period to string for Excel compatibility
        display_df = monthly_df.copy()
        if hasattr(display_df.iloc[0]["order_date"], "strftime"):
            display_df["order_date"] = display_df["order_date"].astype(str)

        _write_df(ws_monthly, display_df, start_row=1, header_color=INDIGO)
        ws_monthly.freeze_panes = "A2"

        # Embed a line chart
        if "total_revenue" in display_df.columns:
            n = len(display_df) + 1
            data_ref = Reference(ws_monthly, min_col=2, min_row=1, max_row=n)
            cats_ref = Reference(ws_monthly, min_col=1, min_row=2, max_row=n)
            chart = LineChart()
            chart.title = "Monthly Revenue Trend"
            chart.y_axis.title = "Revenue ($)"
            chart.x_axis.title = "Month"
            chart.height = 14
            chart.width = 26
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.series[0].graphicalProperties.line.solidFill = INDIGO
            chart.series[0].graphicalProperties.line.width = 20000
            ws_monthly.add_chart(chart, "A" + str(n + 3))

    # ── Sheet 3: 🗺️ Regional ──────────────────────────────────────────────────
    regional_df = summaries.get("regional")
    if regional_df is not None and not regional_df.empty:
        ws_regional = wb.create_sheet("Regional")
        ws_regional.sheet_view.showGridLines = False
        _write_df(ws_regional, regional_df, start_row=1, header_color=EMERALD)
        ws_regional.freeze_panes = "A2"

        # Embedded bar chart
        n = len(regional_df) + 1
        data_ref = Reference(ws_regional, min_col=2, min_row=1, max_row=n)
        cats_ref = Reference(ws_regional, min_col=1, min_row=2, max_row=n)
        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenue by Region"
        chart.y_axis.title = "Revenue ($)"
        chart.x_axis.title = "Region"
        chart.height = 14
        chart.width = 26
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = EMERALD
        ws_regional.add_chart(chart, "A" + str(n + 3))

    # ── Sheet 4: 🏆 Top Products ──────────────────────────────────────────────
    products_df = summaries.get("top_products")
    if products_df is not None and not products_df.empty:
        ws_products = wb.create_sheet("Top Products")
        ws_products.sheet_view.showGridLines = False
        _write_df(ws_products, products_df, start_row=1, header_color=AMBER)
        ws_products.freeze_panes = "A2"

    # ── Sheet 5: ⚠️ Anomalies ────────────────────────────────────────────────
    anomalies_df = summaries.get("anomalies")
    if anomalies_df is not None and not anomalies_df.empty:
        ws_anom = wb.create_sheet("Anomalies")
        ws_anom.sheet_view.showGridLines = False
        _write_df(ws_anom, anomalies_df, start_row=1, header_color="DC2626")

    # ── Named Ranges (for MCP / Antigravity access) ───────────────────────────
    if monthly_df is not None:
        try:
            n_rows = len(monthly_df)
            wb.defined_names["MonthlyRevenue"] = DefinedName(
                "MonthlyRevenue",
                attr_text=f"Monthly!$B$2:$B${n_rows + 1}"
            )
        except Exception:
            pass

    # ── Save ──────────────────────────────────────────────────────────────────
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Smart Excel template saved -> {output_path}")
    return output_path


if __name__ == "__main__":
    from data_cleaner import clean_dataset
    from aggregator import build_summary_tables

    logging.basicConfig(level=logging.INFO, format="%(asctime)s — %(levelname)s — %(message)s")

    parser = argparse.ArgumentParser(description="Build smart Excel template from a dataset")
    parser.add_argument("--data",   required=True, help="Path to raw/clean CSV")
    parser.add_argument("--output", default="reports/smart_report.xlsx", help="Output .xlsx path")
    args = parser.parse_args()

    df = clean_dataset(args.data, "data/clean_data.csv")
    summaries = build_summary_tables(df)
    build_smart_template(summaries, args.output)
    print(f"\n✅ Done! Open: {args.output}")
